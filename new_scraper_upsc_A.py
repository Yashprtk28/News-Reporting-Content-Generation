# -----------------------------------------------------------
# IMPORTS
# -----------------------------------------------------------

import feedparser                 # Parse RSS feeds
import datetime                   # Date & timestamp handling
import asyncio                    # For async Gemini API calls
import os                         # File/folder checks
import schedule                   # Daily scheduler
import time                       # To keep scheduler loop running
import google.generativeai as genai  # Gemini API
from pytrends.request import TrendReq  # Google Trends
from openpyxl import Workbook, load_workbook             # Excel handling
from openpyxl.styles import Font, PatternFill, Alignment # Excel formatting
from openpyxl.utils import get_column_letter             # Auto column width

# -----------------------------------------------------------
# CONFIGURATION
# -----------------------------------------------------------

GEMINI_API_KEY = "AIzaSyC5ejDB8qf9RYaOFF8ty3CpIMZ-mzxT6m8"  # Replace with your key locally
OUTPUT_EXCEL = "trendy_news_summary.xlsx"    # Excel file name

# RSS feeds dictionary: source name -> feed URL
rss_sources = {
    "TheHindu": "https://www.thehindu.com/news/national/rssfeed/",
    "IndianExpress": "https://indianexpress.com/section/india/feed/",
    "GKToday": "https://www.gktoday.in/current-affairs/feed/"
}

# Google Trends setup
pytrends = TrendReq(hl='en-US', tz=330)  # IST timezone

# -----------------------------------------------------------
# SETUP GEMINI CLIENT
# -----------------------------------------------------------

genai.configure(api_key=GEMINI_API_KEY)  # Authenticate Gemini API
gemini_model = genai.GenerativeModel("models/gemini-2.0-flash")  # Initialize model

# -----------------------------------------------------------
# PROMPT BUILDER
# -----------------------------------------------------------

def build_prompt(headline):
    """
    Builds Gemini prompt for a news headline.
    Generates summary, tags, reporter explanation, and 30-second hook.
    """
    return f"""
You are a News Specialist.

Task:
1. Summarize the news headline in 2-3 lines in Hinglish.
2. Suggest relevant tags for the news (general categories like Politics, Economy, Technology, Environment, etc.).
3. Create a reporter-style explanation in Hinglish:
   - Explain background of the issue.
   - Explain key terms, abbreviations, and organisations.
   - Tone should be like a news reporter giving context.

4. Create a crisp 30-second hook-based script for social media:
   - Start with a strong hook in Hinglish.
   - Explain the issue in simple terms.
   - Keep it engaging, short, and clear.

Format:

Summary: <summary>
Tags: <tags>

Reporter_Explanation:
<content>

Hook_30s:
<content>

Headline:
{headline}
"""

# -----------------------------------------------------------
# GEMINI PROCESSOR
# -----------------------------------------------------------

def summarize_and_tag(headline):
    """
    Sends headline to Gemini, parses response into:
    - summary
    - tags
    - reporter explanation
    - 30-second hook
    """
    try:
        response = gemini_model.generate_content(build_prompt(headline))
        lines = response.text.strip().split("\n")
        summary, tags, reporter_expl, hook_30s = "", "", "", ""
        capture_section = None
        for line in lines:
            line = line.strip()
            if line.lower().startswith("summary:"):
                summary = line.split(":",1)[1].strip()
                capture_section = None
            elif line.lower().startswith("tags:"):
                tags = line.split(":",1)[1].strip()
                capture_section = None
            elif line.lower().startswith("reporter_explanation:"):
                capture_section = "reporter"
                reporter_expl = ""
            elif line.lower().startswith("hook_30s:"):
                capture_section = "hook"
                hook_30s = ""
            else:
                if capture_section == "reporter":
                    reporter_expl += line + "\n"
                elif capture_section == "hook":
                    hook_30s += line + "\n"
        return summary, tags, reporter_expl.strip(), hook_30s.strip()
    except Exception as e:
        print("❌ Gemini error:", e)
        return "", "", "", ""

# -----------------------------------------------------------
# DUPLICATE DETECTION
# -----------------------------------------------------------

def load_existing_headlines(filename):
    """
    Reads all sheets in Excel and collects headlines to prevent duplicates
    """
    if not os.path.exists(filename):
        return set()
    wb = load_workbook(filename)
    headlines = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                headlines.add(row[0].strip())
    return headlines

# -----------------------------------------------------------
# GOOGLE TRENDS HELPER
# -----------------------------------------------------------

def extract_keywords(title):
    """
    Simple keyword extraction: split title into words and return first 3-4
    """
    words = [w.strip(".,!?:;") for w in title.split()]
    return words[:4]  # first 4 words

def is_trending(title):
    """
    Checks if headline is trending using Google Trends interest over time
    Returns True if any keyword shows recent high interest
    """
    try:
        keywords = extract_keywords(title)
        pytrends.build_payload(keywords, timeframe='now 7-d', geo='IN')
        df = pytrends.interest_over_time()
        if df.empty:
            return False
        max_interest = df[keywords].max().max()
        return max_interest > 20  # Threshold for "trending"
    except Exception as e:
        print("❌ Google Trends error:", e)
        return False

# -----------------------------------------------------------
# EXCEL FORMATTING
# -----------------------------------------------------------

def format_sheet(ws):
    """
    Formats header row: bold, white font, blue background, center alignment
    Auto adjusts column widths
    """
    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    bold_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 30

# -----------------------------------------------------------
# SAVE TO EXCEL
# -----------------------------------------------------------

def save_to_excel(source_name, all_articles, filename=OUTPUT_EXCEL):
    """
    Saves articles to a sheet named <date>_<source>
    Includes Trend Status column
    """
    today = datetime.date.today().isoformat()
    sheet_name = f"{today}_{source_name}"

    if not os.path.exists(filename):
        wb = Workbook()
        wb.remove(wb.active)
    else:
        wb = load_workbook(filename)

    ws = wb.create_sheet(sheet_name)

    headers = [
        "Title",
        "Link",
        "Published Date",
        "Summary",
        "Tags",
        "Reporter Explanation (Hinglish)",
        "30-Second Hook Script (Hinglish)",
        "Trend Status"
    ]
    ws.append(headers)

    for a in all_articles:
        ws.append([
            a["Title"],
            a["Link"],
            a["Published Date"],
            a["Summary"],
            a["Tags"],
            a["Reporter_Explanation"],
            a["Hook_30s"],
            a["Trend_Status"]
        ])

    format_sheet(ws)
    wb.save(filename)
    print(f"✅ Saved → Sheet: {sheet_name}")

# -----------------------------------------------------------
# MAIN NEWS PIPELINE
# -----------------------------------------------------------

async def run_news_pipeline():
    """
    Main pipeline:
    - Fetch news
    - Check trending status via keyword interest
    - Generate reporter content + 30s hook
    - Save to Excel
    """
    print("\n🚀 Running Trendy News Automation...\n")
    existing_headlines = load_existing_headlines(OUTPUT_EXCEL)

    for source, url in rss_sources.items():
        print(f"\n📌 Fetching: {source}")
        feed = feedparser.parse(url)
        articles = []
        count = 0

        for entry in feed.entries:
            title = entry.title.strip()
            if title in existing_headlines:
                print(f"⚠️ Skipping duplicate: {title}")
                continue
            if count >= 3:  # limit 3 per source
                break

            trending = is_trending(title)
            trend_status = "Trending 🔥" if trending else "Not Trending"

            if trending:
                print(f"📰 Trending News: {title}")
                summary, tags, reporter_expl, hook_30s = summarize_and_tag(title)
            else:
                print(f"ℹ️ Not trending: {title}")
                summary, tags, reporter_expl, hook_30s = "", "", "", ""

            articles.append({
                "Title": title,
                "Link": entry.link,
                "Published Date": entry.get("published", datetime.datetime.now().isoformat()),
                "Summary": summary,
                "Tags": tags,
                "Reporter_Explanation": reporter_expl,
                "Hook_30s": hook_30s,
                "Trend_Status": trend_status
            })

            count += 1

        if articles:
            save_to_excel(source, articles)
        else:
            print(f"ℹ️ No new articles for {source}")

    print("\n🎉 Pipeline Completed!\n")

# -----------------------------------------------------------
# DAILY SCHEDULER
# -----------------------------------------------------------

def start_daily_scheduler(run_time="08:00"):
    """
    Schedule pipeline to run daily at specified time
    """
    schedule.every().day.at(run_time).do(lambda: asyncio.run(run_news_pipeline()))
    print(f"⏳ Scheduler active — will run daily at {run_time}")
    while True:
        schedule.run_pending()
        time.sleep(1)

# -----------------------------------------------------------
# MAIN ENTRY POINT
# -----------------------------------------------------------

if __name__ == "__main__":
    asyncio.run(run_news_pipeline())       # Run pipeline immediately
    # start_daily_scheduler("07:30")       # Uncomment to schedule daily
